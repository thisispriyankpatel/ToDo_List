from typing import Any, Dict
from django.forms.models import BaseModelForm
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.views.generic.list  import ListView
from django.views.generic.detail  import DetailView
from django.views.generic.edit import CreateView, UpdateView, DeleteView, FormView
from django.urls import reverse_lazy
from openpyxl import Workbook
from django.contrib.auth.views import LoginView

from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth import login
from django.db import transaction
from django.views import View
from .models import Task
from .forms import PositionForm


class CustomLoginView(LoginView):
    template_name = 'Base/login.html'
    fields = '__all__'
    redirect_authenticated_user = True

    def get_success_url(self):
        return reverse_lazy('tasks')

class RegisterPage(FormView):
    template_name = 'Base/register.html'
    form_class = UserCreationForm
    redirect_authenticated_user = True
    success_url = reverse_lazy('tasks')


    def form_valid(self, form):
        user = form.save()
        if user is not None:
            login(self.request, user)
        return super().form_valid(form)
    def get(self, *args, **kwargs):
        if self.request.user.is_authenticated:
            return redirect('task') 
        return super(RegisterPage,self).get(*args, **kwargs)


class TaskList(LoginRequiredMixin, ListView):
    model = Task 
    context_object_name = 'tasks'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['tasks'] = context['tasks'].filter(user=self.request.user)
        context['count'] = context['tasks'].filter(complete=False).count()     
        
        search_input = self.request.GET.get('search-area') or ''
        if search_input:
            context['tasks'] = context['tasks'].filter(title__startswith=search_input)

        context['search_input'] = search_input   


        return context
    
class  Downlaod_Task_List(LoginRequiredMixin, ListView):
    model = Task 
    context_object_name = 'tasks', 'task'

    def get_queryset(self):
        queryset = super().get_queryset()
        # Filter the queryset based on the currently logged-in user
        queryset = queryset.filter(user=self.request.user)
        return queryset

    def get(self,request,*args, **kwargs):
        queryset = self.get_queryset()
        workbook = Workbook()
        sheet = workbook.active
        sheet.cell(row=1, column=1, value='Title')
        sheet.cell(row=1, column=2, value='Description')

        for index, task in enumerate(queryset, start=2):
            sheet.cell(row=index, column=1, value=task.title)
            sheet.cell(row=index, column=2, value=task.description)
            
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=mydata.xlsx'
        workbook.save(response)

        return response





class TaskDetail(LoginRequiredMixin, DetailView):
    model = Task
    context_object_name = 'task'
    template_name = 'Base/task.html'

class TaskCreate(LoginRequiredMixin, CreateView):
    model = Task
    fields = ['title', 'description','complete']
    success_url = reverse_lazy('tasks')


    def form_valid(self, form):
        form.instance.user = self.request.user
        return super(TaskCreate,self).form_valid(form)







class TaskUpdate(LoginRequiredMixin, UpdateView):
    model = Task
    fields = ['title', 'description','complete']
    success_url = reverse_lazy('tasks')

class DeleteView(LoginRequiredMixin, DeleteView):
     model = Task
     context_object_name = 'task'
     success_url = reverse_lazy('tasks')

class TaskReorder(View):
    def post(self, request):
        form = PositionForm(request.POST)

        if form.is_valid():
            positionList = form.cleaned_data["position"].split(',')

            with transaction.atomic():
                self.request.user.set_task_order(positionList)

        return redirect(reverse_lazy('tasks'))
    
